import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import config
import utils

HEADERS = [
    "Date", "Punch In", "Lunch Start", "Lunch End", "Punch Out",
    "Comment", "Ad-hoc Hrs", "Ad-hoc Note", "Total Hours",
]
COL_WIDTHS = [12, 10, 12, 10, 10, 30, 11, 30, 12]


class ExcelBackend:
    def __init__(self, path):
        self.path = path
        self._load()

    def _load(self):
        try:
            self.wb = openpyxl.load_workbook(self.path, data_only=True)
        except (FileNotFoundError, Exception):
            # New workbook — default "Sheet" stays until first write_day()
            self.wb = openpyxl.Workbook()

    def _get_or_create_sheet(self, date):
        name = utils.sheet_name_for_date(date)
        # Drop the placeholder default sheet on first real write
        if "Sheet" in self.wb.sheetnames and name not in self.wb.sheetnames:
            del self.wb["Sheet"]
        if name not in self.wb.sheetnames:
            ws = self.wb.create_sheet(name)
            self._write_header(ws)
            self._sort_sheets()
        return self.wb[name]

    def _write_header(self, ws):
        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF")
        hdr_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        for col, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        for col, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    def _sort_sheets(self):
        def sheet_key(name):
            try:
                d = datetime.datetime.strptime(name, "%b %Y")
                return (d.year, d.month)
            except ValueError:
                return (9999, 99)
        self.wb._sheets.sort(key=lambda ws: sheet_key(ws.title))

    def _normalize_date(self, cell_val):
        if cell_val is None:
            return None
        if isinstance(cell_val, datetime.datetime):
            return cell_val.date()
        if isinstance(cell_val, datetime.date):
            return cell_val
        if isinstance(cell_val, str):
            try:
                return datetime.date.fromisoformat(cell_val)
            except ValueError:
                return None
        return None

    # ── Multi-row helpers ───────────────────────────────────────────────────

    def _rows_for_date(self, ws, date):
        """Return sorted list of row indices that belong to date."""
        rows = []
        for row in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=config.COL_DATE).value
            if self._normalize_date(cell_val) == date:
                rows.append(row)
        return sorted(rows)

    def _find_insert_position(self, ws, date):
        """Return the row index at which new rows for date should be inserted.
        That is: the first row whose date is strictly greater than target date,
        or max_row+1 if none (append)."""
        for row in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=config.COL_DATE).value
            d = self._normalize_date(cell_val)
            if d is not None and d > date:
                return row
        return ws.max_row + 1

    def _session_from_row(self, ws, row):
        """Extract punch-time fields from a row as a dict with string values."""
        def g(col):
            v = ws.cell(row=row, column=col).value
            return str(v).strip() if v is not None else ""

        return {
            "punch_in":    g(config.COL_PUNCH_IN),
            "lunch_start": g(config.COL_LUNCH_START),
            "lunch_end":   g(config.COL_LUNCH_END),
            "punch_out":   g(config.COL_PUNCH_OUT),
        }

    def _empty_day(self, date=None):
        return {
            "date":         date,
            "sessions":     [],
            "comment":      "",
            "punch_comment": "",
            "adhoc_hours":  None,
            "adhoc_note":   "",
            "total_hours":  None,
            # Flat convenience fields
            "punch_in":    "",
            "punch_out":   "",
            "lunch_start": "",
            "lunch_end":   "",
        }

    def _day_dict_from_rows(self, ws, rows, date):
        """Build the full day dict from a list of row indices for that date."""
        if not rows:
            return self._empty_day(date)

        sessions = [self._session_from_row(ws, r) for r in rows]

        # Day-level fields are on the first row
        def g(col):
            v = ws.cell(row=rows[0], column=col).value
            return v

        adhoc_raw = g(config.COL_ADHOC_HRS)
        total_raw = g(config.COL_TOTAL_HRS)
        comment   = str(g(config.COL_PUNCH_COMMENT) or "").strip()
        adhoc     = float(adhoc_raw) if adhoc_raw is not None else None
        total     = float(total_raw) if total_raw is not None else None

        first = sessions[0]
        last  = sessions[-1]

        return {
            "date":          date,
            "sessions":      sessions,
            "comment":       comment,
            "punch_comment": comment,
            "adhoc_hours":   adhoc,
            "adhoc_note":    str(g(config.COL_ADHOC_NOTE) or "").strip(),
            "total_hours":   total,
            # Flat convenience fields
            "punch_in":    first.get("punch_in", ""),
            "punch_out":   last.get("punch_out", ""),
            "lunch_start": first.get("lunch_start", ""),
            "lunch_end":   first.get("lunch_end", ""),
        }

    # ── Public API ──────────────────────────────────────────────────────────

    def read_day(self, date):
        name = utils.sheet_name_for_date(date)
        if name not in self.wb.sheetnames:
            return self._empty_day(date)
        ws   = self.wb[name]
        rows = self._rows_for_date(ws, date)
        if not rows:
            return self._empty_day(date)
        return self._day_dict_from_rows(ws, rows, date)

    def write_day(self, date, data):
        ws = self._get_or_create_sheet(date)

        # 1. Delete existing rows for this date (bottom-to-top for stable indices)
        existing = self._rows_for_date(ws, date)
        for row in reversed(existing):
            ws.delete_rows(row)

        # 2. Find insertion position
        insert_at = self._find_insert_position(ws, date)

        # 3. Build the session list to write
        sessions = data.get("sessions") or []
        if not sessions:
            # Write a single blank row to maintain the date entry
            sessions = [{"punch_in": "", "lunch_start": "", "lunch_end": "", "punch_out": ""}]

        # 4. Insert session rows
        n = len(sessions)
        if n > 1:
            ws.insert_rows(insert_at, amount=n)
        elif insert_at > ws.max_row:
            # Appending — no insert needed, row will be created by cell writes
            pass
        else:
            ws.insert_rows(insert_at, amount=1)

        comment   = data.get("punch_comment") or data.get("comment") or ""
        adhoc     = data.get("adhoc_hours")
        adhoc_note= data.get("adhoc_note") or ""

        for idx, session in enumerate(sessions):
            row = insert_at + idx
            ws.cell(row=row, column=config.COL_DATE).value = date

            def sv(col, val):
                ws.cell(row=row, column=col).value = val or None

            sv(config.COL_PUNCH_IN,    session.get("punch_in"))
            sv(config.COL_LUNCH_START, session.get("lunch_start"))
            sv(config.COL_LUNCH_END,   session.get("lunch_end"))
            sv(config.COL_PUNCH_OUT,   session.get("punch_out"))

            if idx == 0:
                # Day-level fields only on first row
                sv(config.COL_PUNCH_COMMENT, comment)
                ws.cell(row=row, column=config.COL_ADHOC_HRS).value = adhoc or None
                sv(config.COL_ADHOC_NOTE, adhoc_note)
            else:
                # Subsequent rows — leave day-level fields blank
                ws.cell(row=row, column=config.COL_PUNCH_COMMENT).value = None
                ws.cell(row=row, column=config.COL_ADHOC_HRS).value     = None
                ws.cell(row=row, column=config.COL_ADHOC_NOTE).value    = None
                ws.cell(row=row, column=config.COL_TOTAL_HRS).value     = None

        # 5. Compute and write total on first row only
        total = utils.compute_day_hours(sessions, adhoc)
        ws.cell(row=insert_at, column=config.COL_TOTAL_HRS).value = total

        try:
            self.wb.save(self.path)
        except PermissionError:
            raise PermissionError(
                "The Excel file is open in another program. Close it first."
            )

    def read_month(self, year, month):
        name = utils.sheet_name_for_date(datetime.date(year, month, 1))
        if name not in self.wb.sheetnames:
            return []
        ws = self.wb[name]

        # Collect all rows grouped by date (preserving order)
        date_to_rows = {}
        date_order   = []
        for row in range(2, ws.max_row + 1):
            date = self._normalize_date(ws.cell(row=row, column=config.COL_DATE).value)
            if date is None:
                continue
            key = date.isoformat()
            if key not in date_to_rows:
                date_to_rows[key] = (date, [])
                date_order.append(key)
            date_to_rows[key][1].append(row)

        result = []
        for key in date_order:
            date, rows = date_to_rows[key]
            result.append(self._day_dict_from_rows(ws, rows, date))
        return result

    def read_all(self):
        result = []
        for name in self.wb.sheetnames:
            try:
                d = datetime.datetime.strptime(name, "%b %Y")
                result.extend(self.read_month(d.year, d.month))
            except ValueError:
                continue
        return result

    def get_months_with_data(self):
        months = []
        for name in self.wb.sheetnames:
            try:
                d = datetime.datetime.strptime(name, "%b %Y")
                months.append((d.year, d.month))
            except ValueError:
                continue
        return months

    def change_path(self, new_path):
        self.wb.save(new_path)
        self.path = new_path
        self._load()
